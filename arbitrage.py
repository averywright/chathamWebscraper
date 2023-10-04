import requests
import xlsxwriter
import pandas as pd
import numpy as np
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill, numbers
"""
ws = xlsxwriter.Workbook('bets.xlsx')

API_KEY = 'e2e5243d2c7087b8e54d7df4451a39e5'

SPORT = 'upcoming' # use the sport_key from the /sports endpoint below, or use 'upcoming' to see the next 8 games across all sports

REGIONS = 'us' # uk | us | eu | au. Multiple can be specified if comma delimited

MARKETS = 'h2h' # h2h | spreads | totals. Multiple can be specified if comma delimited

ODDS_FORMAT = 'decimal' # decimal | american

DATE_FORMAT = 'iso' # iso | unix

BET_SIZE = 100

odds_response = requests.get(
    f'https://api.the-odds-api.com/v4/sports/{SPORT}/odds',
    params={
        'api_key': API_KEY,
        'regions': REGIONS,
        'markets': MARKETS,
        'oddsFormat': ODDS_FORMAT,
        'dateFormat': DATE_FORMAT,
    }
).json()

BOOKMAKER_INDEX = 0
NAME_INDEX = 1
ODDS_INDEX = 2
FIRST = 0


class Event:
    def __init__(self, data):
        self.data = data
        self.sport_key = data['sport_key']
        self.id = data['id']
    
    def find_best_odds(self):
        # number of possible outcomes for a sporting event
        num_outcomes = len(self.data['bookmakers'][FIRST]['markets'][FIRST]['outcomes'])
        self.num_outcomes = num_outcomes
        
        # finding the best odds for each outcome in each event
        best_odds = [[None, None, float('-inf')] for _ in range(num_outcomes)]
        # [Bookmaker, Name, Price]
        
        bookmakers = event.data['bookmakers']
        for index, bookmaker in enumerate(bookmakers):
            
            # determing the odds offered by each bookmaker
            for outcome in range(num_outcomes):
                
                # determining if any of the bookmaker odds are better than the current best odds
                bookmaker_odds = float(bookmaker['markets'][FIRST]['outcomes'][outcome]['price'])
                current_best_odds = best_odds[outcome][ODDS_INDEX]
                
                if bookmaker_odds > current_best_odds:
                    best_odds[outcome][BOOKMAKER_INDEX] = bookmaker['title']
                    best_odds[outcome][NAME_INDEX] = bookmaker['markets'][FIRST]['outcomes'][outcome]['name']
                    best_odds[outcome][ODDS_INDEX] = bookmaker_odds
        
        self.best_odds = best_odds
        return best_odds
    
    def arbitrage(self):
        total_arbitrage_percentage = 0
        for odds in self.best_odds:
            total_arbitrage_percentage += (1.0 / odds[ODDS_INDEX])
        
        self.total_arbitrage_percentage = total_arbitrage_percentage
        self.expected_earnings = (BET_SIZE / total_arbitrage_percentage) - BET_SIZE
        
        # if the sum of the reciprocals of the odds is less than 1, there is opportunity for arbitrage
        if total_arbitrage_percentage < 1:
            return True
        return False
    
    # converts decimal/European best odds to American best odds
    def convert_decimal_to_american(self):
        best_odds = self.best_odds
        for odds in best_odds:
            decimal = odds[ODDS_INDEX]
            if decimal >= 2:
                american = (decimal - 1) * 100
            elif decimal < 2:
                american = -100 / (decimal - 1)
            odds[ODDS_INDEX] = round(american, 2)
        return best_odds
    
    def calculate_arbitrage_bets(self):
        bet_amounts = []
        for outcome in range(self.num_outcomes):
            individual_arbitrage_percentage = 1 / self.best_odds[outcome][ODDS_INDEX]
            bet_amount = (BET_SIZE * individual_arbitrage_percentage) / self.total_arbitrage_percentage
            bet_amounts.append(round(bet_amount, 2))
        
        self.bet_amounts = bet_amounts
        return bet_amounts


events = []
for data in odds_response:
    events.append(Event(data))
    # print(data)
    # print()

arbitrage_events = []
for event in events:
    best_odds = event.find_best_odds()
    if event.arbitrage():
        arbitrage_events.append(event)

for event in arbitrage_events:
    print(event)
    event.calculate_arbitrage_bets()
    event.convert_decimal_to_american()
    
MAX_OUTCOMES = max([event.num_outcomes for event in arbitrage_events])
ARBITRAGE_EVENTS_COUNT = len(arbitrage_events)

my_columns = ['ID', 'Sport Key', 'Expected Earnings'] + list(np.array([[f'Bookmaker #{outcome}', f'Name #{outcome}', f'Odds #{outcome}', f'Amount to Buy #{outcome}'] for outcome in range(1, MAX_OUTCOMES + 1)]).flatten())
dataframe = pd.DataFrame(columns=my_columns)

for event in arbitrage_events:
    # print(event.best_odds)
    row = []
    row.append(event.id)
    row.append(event.sport_key)
    row.append(round(event.expected_earnings, 2))
    for index, outcome in enumerate(event.best_odds):
        row.append(outcome[BOOKMAKER_INDEX])
        row.append(outcome[NAME_INDEX])
        row.append(outcome[ODDS_INDEX])
        row.append(event.bet_amounts[index])
    while len(row) < len(dataframe.columns):
        row.append('N/A')
    dataframe.loc[len(dataframe.index)] = row
    print(row)


"""
import re

API_KEY = 'e2e5243d2c7087b8e54d7df4451a39e5'

SPORT = 'upcoming' # use the sport_key from the /sports endpoint below, or use 'upcoming' to see the next 8 games across all sports

REGIONS = 'us' # uk | us | eu | au. Multiple can be specified if comma delimited

MARKETS = 'h2h' # h2h | spreads | totals. Multiple can be specified if comma delimited

ODDS_FORMAT = 'decimal' # decimal | american

DATE_FORMAT = 'iso' # iso | unix

BET_SIZE = 100

odds_response = requests.get(
    f'https://api.the-odds-api.com/v4/sports/{SPORT}/odds',
    params={
        'api_key': 'e2e5243d2c7087b8e54d7df4451a39e5',
        'regions': REGIONS,
        'markets': MARKETS,
        'oddsFormat': ODDS_FORMAT,
        'dateFormat': DATE_FORMAT,
    }
).json()


arb = {}
counter = 0
for i in odds_response:
    for j in i['bookmakers']:
        for k in j:
            counter += 1
            if (counter - 1) % 4 == 0:
                arb[j[k]] = []

for i in odds_response:
    for j in i['bookmakers']:
        for k in j:
            counter += 1
            if (counter - 1) % 4 == 0:
                temp = j[k]
            if counter % 4 == 0:
                for m in j[k][0]['outcomes']:
                    arb[temp].append({m['name']:m['price']})


print(arb)
counter = 0
newValues = {}
for i in arb:
    if i == 'bovada' or i == 'mybookieag' or i == 'betus' or i == 'betonlineag' or i == 'lowvig':
        counter += 1
        if counter == 1:
            count = 0
            for j in arb[i]:
                count += 1
                for k in j:
                    if k == 'Draw':
                        for s in arb[i][count-2]:
                            for t in arb[i][count - 3]:
                                newValues[str(s) + str(t) + ' Draw'] = j[k]
                    else:
                        newValues[k] = j[k]
                    
        else:
            count = 0
            for j in arb[i]:
                count += 1
                for k in j:
                    try:
                        if k == 'Draw':
                            for s in arb[i][count - 2]:
                                for t in arb[i][count - 3]:
                                    if j[k][0] > newValues[str(s) + str(t) + ' Draw']:
                                        newValues[str(s) + str(t) + ' Draw'] = [j[k], i]
                        elif j[k][0] > newValues[k]:
                            newValues[k] = [j[k], i]
                            
                    except:
                        if k == 'Draw':
                            for s in arb[i][count - 2]:
                                for t in arb[i][count - 3]:
                                    newValues[str(s) + str(t) + ' Draw'] = [j[k],i]
                        else:
                            newValues[k] = [j[k], i]

print(newValues)
counter = 0
total = 0
listed = []
books = []
for i in newValues:
    counter += 1
    total += 1/newValues[i][0]
    books.append(newValues[i])
    listed.append(i)
    if counter == 3:
        print(books)
        print(listed)
        print(total)
        listed = []
        books = []
        total = 0
        counter = 0
